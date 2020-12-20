import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import win32com.client
import os
import poisson_distribution

# 엑셀을 제작하는 클래스로, 주로 사이트에서 과거의 기록들을 가져와 조작하여 엑셀에 작성. 직접 제작한 알고리즘으로부터 도출된 결과도 작성하며
# 과거의 기록들에 정상성을 확인하고자 푸아송 분포를 활용함
class Make_Excel():

    # 팀 이름을 전달해주어서 history 를 출력시키고, 임의로 스탯 설정한 것도 받아서 출력시켜야함.
    def __init__(self, TEAM1, TEAM2, Expected_result):
        self.TEAM1 = TEAM1
        self.TEAM2 = TEAM2
        self.Expected_result = Expected_result

    def start(self):
        # team 들의 이름을 replace문을 이용해 재설정 해주었는데, 검색 시에 첫번째 팀에 공백이 있는 경우엔 -를 넣고 있고,
        # 두번째 팀의 이름에 -가 들어가 있는 경우 공백으로 처리하고 있었기 때문
        self.TEAM1 = self.TEAM1.replace(" ", "-")
        self.TEAM2 = self.TEAM2.replace("-", " ")
        print("---------------")
        print(self.TEAM1)
        print(self.TEAM2)
        # team1의 이름을 넣고, 해당 url을 직접 들어가보면 과거 데이터들을 확인할 수 있음
        url = 'https://www.11v11.com/teams/' + self.TEAM1 + '/tab/opposingTeams/opposition/' + self.TEAM2 + '/'

        # 엑셀 파일에 쓰기
        write_wb = Workbook()
        write_ws = write_wb.active

        response = requests.get(url)
        soup = BeautifulSoup(response.content, 'lxml')

        # 홈페이지 엤는 table 을 찾는 과정. 저 table 이름이 홈페이지에 있는 선수들 스탯 창
        table = soup.find("table",
                          {"class": "sortable"})
        data = []
        for a in table.find_all("tr"):
            for b in a.find_all("td"):
                data.append(b.get_text())
        i = 0
        result_list = []
        result = []
        count_play = 0
        count_win = 0
        count_lose = 0
        count_draw = 0
        team1_score = 0
        team2_score = 0
        for i in range(0, int(len(data) / 5)):

            # 해당 if 문을 삼입한 이유는, 해당 홈페이지에서 아직 경기가 이루어 지지 않은, 곧 경기가 치뤄질 데이터들을 넣고 있기 때문에
            # 그런 경우를 제하기 위함
            if data[5 * i + 3] == '':
                continue
            # 2015 이전의 자료들을 제외하기 위한 조건문
            if int(data[5 * i].split(" ")[2]) < 2015:
                continue
            else:
                count_play = count_play + 1
                # 사이트에서 뒤죽박죽으로 결과가 저장되어 있어 TEAM1 v TEAM2로 결과를 재정렬함. 팀 명 뿐만 아니라 다른 값들 위치를 바꿔야함
                # 사이트에서 index 별로 [날짜, 매치하는 팀들, 결과(W or L or D), 스코어, 경기종류] 형태로 테이블이 이루어져있음
                # 따라서 5 를 기준으로 새로운 경기들의 결과를 나타냄. 이후에는 각 유의미한 데이터들을 정리하기 위해 파싱함
                if data[5 * i + 1][0] == self.TEAM2[0].capitalize():
                    data[5 * i + 1] = self.TEAM1.title() + " v " + self.TEAM2.title()
                    # 3번째 인덱스는 경기 스코어로, 하이푼을 기준으로 점수를 나누어 주었음. 섞여 있는 탓에 스코어를 뒤바꾸어 주어야 하기 때문
                    j = data[5 * i + 3].split("-")
                    team1_score = team1_score + int(j[1])
                    team2_score = team2_score + int(j[0])
                    data[5 * i + 3] = j[1] + "-" + j[0]
                else:
                    j = data[5 * i + 3].split("-")
                    team1_score = team1_score + int(j[0])
                    team2_score = team2_score + int(j[1])

                # 경기 승률 체크/ 결과에 따라 값들을 가산함
                if data[5 * i + 2] == "W":
                    count_win = count_win + 1
                elif data[5 * i + 2] == "L":
                    count_lose = count_lose + 1
                else:
                    count_draw = count_draw + 1

                # result 리스트에 필요한 정보만을 저장함. 이 경우엔 날짜와 결과, 스코어만을 result 에 저장함
                result.append(data[5 * i + 0])
                result.append(data[5 * i + 2])
                result.append(data[5 * i + 3])

                result_list.append(result)
                result = []
        # 결과들을 간략하게 출력시키는 부분
        for i in result_list:
            print(i)
        print("play: %d win: %d lose %d draw: %d" % (count_play, count_win, count_lose, count_draw))
        print("team1_score_aver: %.3f team2_score_aver: %.3f" % (team1_score / count_play, team2_score / count_play))

        # 현재 cwd 를 출력
        print(os.getcwd())

        # 엑셀의 해당 위치에 결과들을 write
        write_ws['A1'] = self.TEAM1 + "  vs  " + self.TEAM2
        write_ws.append(["Date", "Result", "Score", "", "Average score", "", "Play", "Win", "Lose", "Draw"])

        for i in result_list:
            write_ws.append(i)

        # 엑셀 작성
        write_ws['E1'] = "<History>"
        write_ws['E5'] = "<Average Win>"
        write_ws['E3'] = str(format(team1_score / count_play, ".3f")) + " vs " + str(
            format(team2_score / count_play, ".3f"))
        write_ws['E6'] = format(count_win / count_play * 100, ".3f") + "%"
        write_ws['G3'] = count_play
        write_ws['H3'] = count_win
        write_ws['I3'] = count_lose
        write_ws['J3'] = count_draw
        write_ws['E9'] = "<MY RESULT>"
        write_ws['L1'] = "<POISSON>"
        write_ws['M2'] = self.TEAM1
        write_ws['M3'] = self.TEAM2
        for i in range(0, 10):
            write_ws.cell(row=1, column=i + 14).value = i
        # poisson 에 해당 팀들의 평균 골의 값을 넘긴 뒤, 스코어 중 확률이 가장 높은 것 엑셀에 작성
        hel = poisson_distribution.Poisson(team1_score / count_play, team2_score / count_play)
        [list_1, list_2] = hel.get_poisson_list()

        # 엑셀에 team1과 team2의 poisson 값들을 작성. 아스키와 문자를 이용하여 엑셀에 차례대로 입력시킴
        for i in range(2, 12):
            write_ws.cell(row=2, column=14 + (i - 2)).value = (format(float(list_1[i - 2][1] * 100), ".3f")) + "%"
            write_ws.cell(row=3, column=14 + (i - 2)).value = (format(float(list_2[i - 2][1] * 100), ".3f")) + "%"

        # poisson 에서 최대 확률을 찾음으로써 스코어 결과 도출
        max_1 = 0.0
        max_2 = 0.0

        # 푸아송 분포를 통해 구한 결과들 중 제일 높은 확률을 보인 score 를 구함
        for i in range(0, 10):
            if max_1 <= list_1[i][1]:
                max_1 = list_1[i][1]
                max_1_index = i
            if max_2 <= list_2[i][1]:
                max_2 = list_2[i][1]
                max_2_index = i
        # poisson 으로 구한 결과 엑셀에 표시. 결과는 노란색으로 입히기
        write_ws['L5'] = "<POISSON RESULT>"

        # 엑셀에 예상된 결과를 출력시킴, 액셀에서의 색깔 값을 이용해 설정
        yellowFill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        redFill = PatternFill(start_color='FF4500', end_color='FF4500', fill_type='solid')
        GreenFill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')

        write_ws.cell(row=6, column=12).fill = yellowFill

        # 푸아송 분포로 구한 뒤, 확률이 제일 높은 스코어를 비교하여 승,패를 파악한 뒤 엑셀에 출력함.
        if max_1_index > max_2_index:
            write_ws['L6'] = str(max_1_index) + " vs " + str(max_2_index) + " (W) "
        elif max_1_index == max_2_index:
            write_ws['L6'] = str(max_1_index) + " vs " + str(max_2_index) + " (D) "
        else:
            write_ws['L6'] = str(max_1_index) + " vs " + str(max_2_index) + " (L) "

        write_ws['E6'].fill = GreenFill

        # 직접 예상하는 경기 스코어를 엑셀에 입력함. 경기 결과에 따라 W, L, D를 출력시키기 위한 조건문
        write_ws.cell(row=10, column=5).fill = redFill
        if int(self.Expected_result["goal"][0]) > int(self.Expected_result["goal"][1]):
            write_ws.cell(row=10, column=5).value = self.Expected_result["goal"][0] + " vs " + \
                                                       self.Expected_result["goal"][1] + "(W)"
        elif int(self.Expected_result["goal"][0]) == int(self.Expected_result["goal"][1]):
            write_ws.cell(row=10, column=5).value = self.Expected_result["goal"][0] + " vs " + \
                                                       self.Expected_result["goal"][1] + "(D)"
        else:
            write_ws.cell(row=10, column=5).value = self.Expected_result["goal"][0] + " vs " + \
                                                       self.Expected_result["goal"][1] + "(L)"
        # write 한 값들을 저장하기 위해 해당 위치의 파일에 save
        write_wb.save(os.getcwd() + "/result_list.xlsx")
        write_wb.close()

        # 엑셀 파일을 win32com 을 이용해서 open. Pannel 과 연동하기 위해서 excel 객체를 return
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Workbooks.Open(os.getcwd() + "/result_list.xlsx")

        return excel
